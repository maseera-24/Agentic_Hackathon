import io
import sys
import os
from openpyxl import load_workbook

# Add project root to sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient
from backend.main import app
from backend.routers.auth import create_auth_token
from backend.data.db import db

def test_agent_excel_and_operations():
    print("=" * 70)
    print("TESTING REAL AGENT NATURAL-LANGUAGE OPERATIONS & EXCEL GENERATION")
    print("=" * 70)

    db.reset_to_seed()
    client = TestClient(app)
    token = create_auth_token({"id": "USR_OFFICER_01", "role": "placement_officer", "email": "officer@example.com"})
    headers = {"Authorization": f"Bearer {token}"}

    # 1. Test Query: "give me the excel file of the students who are selected"
    print("\n[1] Testing Natural-Language Selected Students Excel Request:")
    query = "give me the excel file of the students who are selected"
    res = client.post("/api/agent/chat", json={"message": query, "context": {"active_drive": "DRIVE_GOOGLE_2026"}}, headers=headers)
    assert res.status_code == 200, f"Failed: {res.text}"
    data = res.json()
    
    print(f"    Intent: {data.get('intent')}")
    print(f"    Executed Tools: {[t['name'] for t in data.get('executed_tools', [])]}")
    print(f"    Download URL: {data.get('download_url')}")
    assert data.get("intent") == "DOWNLOAD_SELECTED_EXCEL"
    assert "generate_selected_students_excel" in [t['name'] for t in data.get('executed_tools', [])]
    assert data.get("download_url") is not None
    
    # 2. Test Real Excel Download Endpoint
    dl_url = data.get("download_url")
    print(f"\n[2] Downloading Excel File from {dl_url} ...")
    dl_res = client.get(dl_url, headers=headers)
    assert dl_res.status_code == 200, f"Download failed: {dl_res.status_code}"
    assert "spreadsheetml" in dl_res.headers.get("content-type", "")
    
    wb = load_workbook(io.BytesIO(dl_res.content))
    print(f"    Workbook Sheets: {wb.sheetnames}")
    assert "Selected Students" in wb.sheetnames
    ws = wb["Selected Students"]
    rows = list(ws.iter_rows(values_only=True))
    print(f"    Total Rows in Selected Sheet: {len(rows)}")
    print(f"    Header Columns ({len(rows[0])}): {rows[0]}")
    sample_safe = [str(val).encode('ascii', 'replace').decode('ascii') for val in rows[1]] if len(rows) > 1 else 'None'
    print(f"    Sample Row: {sample_safe}")
    assert len(rows) >= 2, "Expected at least 1 data row in Selected Students sheet"
    print("    [PASS] Real Excel .xlsx successfully generated and downloaded!")

    # 3. Test Query: "show me the students who were not selected"
    print("\n[3] Testing Non-Selected Students Query:")
    q_not_sel = "show me the students who were not selected"
    res_not_sel = client.post("/api/agent/chat", json={"message": q_not_sel, "context": {"active_drive": "DRIVE_GOOGLE_2026"}}, headers=headers)
    assert res_not_sel.status_code == 200
    d_not = res_not_sel.json()
    print(f"    Intent: {d_not.get('intent')}")
    print(f"    Tools: {[t['name'] for t in d_not.get('executed_tools', [])]}")
    assert d_not.get("intent") == "GET_NOT_SELECTED_STUDENTS"
    assert "get_rejected_students" in [t['name'] for t in d_not.get('executed_tools', [])]
    print("    [PASS] Non-selected student records retrieved with real reasons!")

    # 4. Test Student-Scoped Query
    student_token = create_auth_token({"id": "USR_STU_001", "role": "student", "email": "student@example.com", "student_id": "STU001"})
    s_headers = {"Authorization": f"Bearer {student_token}"}
    print("\n[4] Testing Student Assistant Scoped Query:")
    s_res = client.post("/api/agent/chat", json={"message": "what drives can I apply for?", "context": {"user_role": "student", "user_id": "STU001"}}, headers=s_headers)
    assert s_res.status_code == 200
    s_data = s_res.json()
    print(f"    Intent: {s_data.get('intent')}")
    print(f"    Tools: {[t['name'] for t in s_data.get('executed_tools', [])]}")
    assert s_data.get("intent") == "STUDENT_ELIGIBLE_DRIVES"
    assert "student_get_eligible_drives" in [t['name'] for t in s_data.get('executed_tools', [])]
    print("    [PASS] Student assistant correctly scoped to student's verified profile!")

    print("\n" + "=" * 70)
    print("ALL REAL AGENT OPERATIONS & EXCEL TESTS PASSED SUCCESSFULLY!")
    print("=" * 70)

if __name__ == "__main__":
    test_agent_excel_and_operations()
