import os
import sys
import io
import re
from openpyxl import load_workbook
from fastapi.testclient import TestClient

# Add project root to sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.main import app
from backend.data.db import db
from backend.routers.auth import create_auth_token

client = TestClient(app)

def test_excel_export_pipeline():
    print("=" * 60)
    print("TESTING REAL EXCEL EXPORT PIPELINE")
    print("=" * 60)

    # 1. Obtain officer token
    officer_user = db.get_user_by_email("officer@example.com")
    assert officer_user is not None, "Officer user not found"
    token = create_auth_token(officer_user)
    headers = {"Authorization": f"Bearer {token}"}
    print(f"[1] Officer authenticated: {officer_user['name']} ({officer_user['role']})")

    # 2. Get active drives
    drives = db.get_drives()
    assert len(drives) > 0, "No drives found in MongoDB"
    stripe_drive = next((d for d in drives if "Stripe" in d.get("company_name", "")), drives[0])
    drive_id = stripe_drive["id"]
    company_name = stripe_drive.get("company_name")
    print(f"[2] Target Drive: {company_name} (ID: {drive_id})")

    # 3. Test Shortlist Export with Bearer Header
    print("\n[3] Testing GET /api/drives/{drive_id}/shortlist/export (Bearer Header)...")
    res1 = client.get(f"/api/drives/{drive_id}/shortlist/export", headers=headers)
    assert res1.status_code == 200, f"Expected 200, got {res1.status_code}: {res1.text}"
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res1.headers["content-type"], f"Invalid content-type: {res1.headers.get('content-type')}"
    disp1 = res1.headers.get("content-disposition", "")
    assert "shortlist_" in disp1 and ".xlsx" in disp1, f"Invalid content-disposition: {disp1}"
    print(f"    [+] Status: 200 OK | Content-Type: {res1.headers['content-type']}")
    print(f"    [+] Content-Disposition: {disp1}")

    # Inspect Workbook
    wb1 = load_workbook(io.BytesIO(res1.content))
    assert "Shortlisted Students" in wb1.sheetnames, f"Missing Shortlisted sheet: {wb1.sheetnames}"
    assert "Rejected Students" in wb1.sheetnames, f"Missing Rejected sheet: {wb1.sheetnames}"

    ws1_short = wb1["Shortlisted Students"]
    short_headers = [cell.value for cell in ws1_short[1]]
    assert "Student ID" in short_headers and "Skill Match" in short_headers, f"Invalid headers: {short_headers}"
    print(f"    [+] Sheet 1 'Shortlisted Students': {ws1_short.max_row - 1} candidate rows, Columns: {short_headers[:5]}...")

    ws1_rej = wb1["Rejected Students"]
    rej_headers = [cell.value for cell in ws1_rej[1]]
    assert "Rejection Reason" in rej_headers, f"Missing Rejection Reason in rejected sheet: {rej_headers}"
    print(f"    [+] Sheet 2 'Rejected Students': {ws1_rej.max_row - 1} candidate rows, Columns: {rej_headers[:5]}...")

    # 4. Test Shortlist Export with Query Token (Direct browser download simulation)
    print("\n[4] Testing GET /api/drives/{drive_id}/shortlist/export?auth={token} (Query Token)...")
    res2 = client.get(f"/api/drives/{drive_id}/shortlist/export?auth={token}")
    assert res2.status_code == 200, f"Expected 200 with query token, got {res2.status_code}"
    print(f"    [+] Status: 200 OK with query token ({len(res2.content)} bytes)")

    # 5. Test Selected Students Export
    print("\n[5] Testing GET /api/drives/{drive_id}/results/export/selected...")
    res_sel = client.get(f"/api/drives/{drive_id}/results/export/selected", headers=headers)
    assert res_sel.status_code == 200
    disp_sel = res_sel.headers.get("content-disposition", "")
    assert "selected_students_" in disp_sel and ".xlsx" in disp_sel
    wb_sel = load_workbook(io.BytesIO(res_sel.content))
    print(f"    [+] Selected Students Workbook: {wb_sel.sheetnames} ({len(res_sel.content)} bytes)")
    print(f"    [+] Filename: {disp_sel}")

    # 6. Test Not Selected Students Export
    print("\n[6] Testing GET /api/drives/{drive_id}/results/export/not-selected...")
    res_not_sel = client.get(f"/api/drives/{drive_id}/results/export/not-selected", headers=headers)
    assert res_not_sel.status_code == 200
    disp_not_sel = res_not_sel.headers.get("content-disposition", "")
    assert "not_selected_students_" in disp_not_sel and ".xlsx" in disp_not_sel
    wb_not_sel = load_workbook(io.BytesIO(res_not_sel.content))
    print(f"    [+] Not Selected Students Workbook: {wb_not_sel.sheetnames} ({len(res_not_sel.content)} bytes)")
    print(f"    [+] Filename: {disp_not_sel}")

    # 7. Test Complete Results Multi-sheet Export
    print("\n[7] Testing GET /api/drives/{drive_id}/results/export/all...")
    res_all = client.get(f"/api/drives/{drive_id}/results/export/all", headers=headers)
    assert res_all.status_code == 200
    disp_all = res_all.headers.get("content-disposition", "")
    assert "complete_results_" in disp_all and ".xlsx" in disp_all
    wb_all = load_workbook(io.BytesIO(res_all.content))
    assert "Executive Summary" in wb_all.sheetnames
    print(f"    [+] Complete Results Multi-Sheet Workbook: {wb_all.sheetnames} ({len(res_all.content)} bytes)")
    print(f"    [+] Filename: {disp_all}")

    print("\n" + "=" * 60)
    print("ALL 4 EXCEL EXPORT ENDPOINTS FULLY VERIFIED & WORKING!")
    print("=" * 60)

if __name__ == "__main__":
    test_excel_export_pipeline()
